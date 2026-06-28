# -*- coding: utf-8 -*-
"""Import price list from Excel into price.js structure."""
import json
import re
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r'c:\Users\user\Downloads\Новая таблица.xlsx')

SUBSECTIONS_COL2 = [
    ('Брекет-системы', 'braces'),
    ('Аппараты и пластинки', 'apparatus'),
    ('Элайнеры и каппы', 'aligners'),
]


def normalize_price(raw: str) -> str:
    s = raw.strip().rstrip('.')
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('ТГ', 'тг').replace(' Тг', ' тг')
    s = re.sub(r'(\d)\s+т\s*$', r'\1 тг', s, flags=re.I)
    s = re.sub(r'(\d)тг', r'\1 тг', s, flags=re.I)
    if re.search(r'\d', s) and 'тг' not in s.lower():
        s += ' тг'
    if s.lower().startswith('от ') and not s.endswith('тг'):
        s += ' тг'
    return s


def split_item(cell):
    if cell is None:
        return None
    text = str(cell).strip().replace('\n', ' ')
    if not text:
        return None
    if text.endswith(':') or re.fullmatch(r'[А-ЯA-Z][А-ЯA-Z\s\-]+:?', text):
        return ('header', text.rstrip(':').strip())

    m = re.match(r'^(.+?)\s*(?:—|–)\s*(.+)$', text)
    if m:
        return ('item', m.group(1).strip().rstrip('.'), normalize_price(m.group(2)))

    m = re.match(r'^(.+?)\s+-\s+(от\s.+|\d.+)$', text, re.I)
    if m:
        return ('item', m.group(1).strip().rstrip('.'), normalize_price(m.group(2)))

    return ('item', text.rstrip('.'), '')


def parse_column(ws, col_idx, skip_headers=None):
    skip_headers = skip_headers or []
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = row[col_idx] if col_idx < len(row) else None
        parsed = split_item(cell)
        if not parsed:
            continue
        if parsed[0] == 'header':
            if any(h.lower() in parsed[1].lower() for h in skip_headers):
                continue
            continue
        items.append((parsed[1], parsed[2]))
    return items


def parse_orthodontics(ws):
    sections = {k: [] for _, k in SUBSECTIONS_COL2}
    current = 'braces'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = row[2] if len(row) > 2 else None
        parsed = split_item(cell)
        if not parsed:
            continue
        if parsed[0] == 'header':
            for prefix, key in SUBSECTIONS_COL2:
                if parsed[1].lower().startswith(prefix.lower()):
                    current = key
                    break
            continue
        sections[current].append((parsed[1], parsed[2]))
    return sections


def parse_surgery_periodont(ws):
    surgery, perio = [], []
    mode = 'surgery'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = row[4] if len(row) > 4 else None
        parsed = split_item(cell)
        if not parsed:
            continue
        if parsed[0] == 'header' and 'ПАРОДОНТ' in parsed[1].upper():
            mode = 'perio'
            continue
        if parsed[0] == 'header':
            continue
        (perio if mode == 'perio' else surgery).append((parsed[1], parsed[2]))
    return surgery, perio


def parse_workbook():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    ortho = parse_orthodontics(ws)
    surgery, perio = parse_surgery_periodont(ws)

    data = {
        'therapy': {'label': 'Терапия', 'items': parse_column(ws, 0)},
        'orthopaed': {'label': 'Ортопедия', 'items': parse_column(ws, 1)},
        'braces': {'label': 'Брекеты', 'items': ortho['braces']},
        'apparatus': {'label': 'Аппараты и пластины', 'items': ortho['apparatus']},
        'aligners': {'label': 'Элайнеры и капы', 'items': ortho['aligners']},
        'hygiene': {'label': 'Гигиена', 'items': parse_column(ws, 3)},
        'periodontology': {'label': 'Пародонтология', 'items': perio},
        'kids': {'label': 'Детская стоматология', 'items': parse_column(ws, 6)},
        'surgery': {'label': 'Хирургия', 'items': surgery},
        'implants': {'label': 'Имплантация', 'items': parse_column(ws, 5)},
        'diagnostics': {'label': 'Диагностика', 'items': parse_column(ws, 7)},
    }
    return {k: v for k, v in data.items() if v['items']}


def js_string(s):
    return json.dumps(s, ensure_ascii=False)


def generate_price_js(data):
    lines = [
        '/* =============================================',
        '   price.js — Price list tabs + search',
        '   Arlan Med Dental Clinic',
        '   ============================================= */',
        '',
        'const priceData = {',
    ]
    for key, cat in data.items():
        lines.append(f"  {key}: {{")
        lines.append(f"    label: {js_string(cat['label'])},")
        lines.append('    items: [')
        for name, price in cat['items']:
            lines.append(f"      [{js_string(name)}, {js_string(price)}],")
        lines.append('    ]')
        lines.append('  },')
    lines.append('};')
    lines.append('')
    return '\n'.join(lines)


if __name__ == '__main__':
    data = parse_workbook()
    out_json = ROOT / '_price_imported.json'
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # price.js
    price_js_path = ROOT / 'js' / 'price.js'
    old_tail = price_js_path.read_text(encoding='utf-8').split('const priceData = {', 1)[1]
    old_tail = old_tail.split('};', 1)[1]
    price_js_path.write_text(generate_price_js(data) + old_tail, encoding='utf-8')

    # price-i18n.js
    total = 0
    lines = [
        '/* =============================================',
        '   price-i18n.js — Price list item translations (RU / KZ / EN)',
        '   Arlan Med Dental Clinic',
        f'   Generated from Excel — {sum(len(v["items"]) for v in data.values())} items',
        '   ============================================= */',
        '',
        'const priceTranslations = {',
    ]
    for lang in ('ru', 'kz', 'en'):
        lines.append(f'  {lang}: {{')
        for cat_key, cat in data.items():
            for i, (name, _price) in enumerate(cat['items']):
                key = f'pr_{cat_key}_{i}'
                lines.append(f'    {key}: {js_string(name)},')
                total += 1 if lang == 'ru' else 0
        lines.append('  },')
    lines.append('};')
    lines.append('')
    lines.append('window.priceTranslations = priceTranslations;')
    lines.append('')
    (ROOT / 'js' / 'price-i18n.js').write_text('\n'.join(lines), encoding='utf-8')

    for k, v in data.items():
        print(k, len(v['items']))
    print('total', sum(len(v['items']) for v in data.values()))
    print('Written js/price.js and js/price-i18n.js')
